"""
importador.py
=============
Importador do Excel: Base.2025.resultados SSO.xlsx
Projeto: Dashboard Comercial Executivo SSO — Etapas 1 e 2

Lê as 7 abas mensais, localiza colunas pelos TEXTOS dos cabeçalhos
(nunca por posição fixa), aplica todas as regras de tratamento e
produz a base analítica normalizada (sem nenhum dado pessoal).

Autor: Antigravity / Google DeepMind
Data:  2026-07-21
"""

import re
import sys
import json
import uuid
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

import openpyxl

# Importar classificador Curva ABC (mesmo diretório src/)
try:
    from curva_abc import converter_qtd_funcionarios, classificar_curva
except ImportError:
    sys.path.insert(0, str(Path(__file__).parent))
    from curva_abc import converter_qtd_funcionarios, classificar_curva

# ---------------------------------------------------------------------------
# CONFIGURAÇÕES
# ---------------------------------------------------------------------------

ARQUIVO_EXCEL = Path(__file__).parent.parent / "Base.2025.resultados SSO.xlsx"
SAIDA_JSON    = Path(__file__).parent.parent / "output" / "base_analitica.json"
LOG_QUALIDADE = Path(__file__).parent.parent / "output" / "qualidade.json"

# Mapeamento aba → (numero_mes, nome_mes, ano)
ABA_PARA_MES: Dict[str, Tuple[int, str, int]] = {
    "Janeiro26":   (1,  "Janeiro",   2026),
    "Fevereiro26": (2,  "Fevereiro", 2026),
    "Março26":     (3,  "Março",     2026),
    "Abril26":     (4,  "Abril",     2026),
    "Maio26":      (5,  "Maio",      2026),
    "Junho26":     (6,  "Junho",     2026),
    "Julho26":     (7,  "Julho",     2026),
}

# Linha de cabeçalhos (1-indexed)
LINHA_HEADER = 4

# Nomes canônicos dos cabeçalhos que queremos capturar
# (compara após strip + upper para resistir a variações)
HEADER_QTD           = "QTD"
HEADER_TIPO_CONTRATO = "TIPO DE CONTRATO"   # primeira ocorrência
HEADER_VALOR_TOTAL   = "VALOR TOTAL"
HEADER_STATUS        = "STATUS"
HEADER_FONTE_LEAD    = "FONTE DO LEAD"
HEADER_VENDEDOR      = "VENDEDOR"
HEADER_QTD_FUNC      = "QUANTIDADE DE FUNCIONARIOS"  # também aceita com Ã

# Aliases para cabeçalhos com acentuação variável encontrados na planilha
HEADER_VENDEDOR_ALIASES = {"VENDEDOR"}
HEADER_FONTE_ALIASES    = {"FONTE DO LEAD"}

# Normalização para comparação de STATUS
STATUS_CONTRATO_FECHADO = "CONTRATO FECHADO"

# ---------------------------------------------------------------------------
# LOGGER
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("importador_sso")


# ---------------------------------------------------------------------------
# FUNÇÕES AUXILIARES
# ---------------------------------------------------------------------------

def _strip(valor: Any) -> str:
    """Remove apenas espaços no início e no final; preserva o restante."""
    if valor is None:
        return ""
    return str(valor).strip()


def _normalizar_status(valor: Any) -> str:
    """Trim + upper para comparação de STATUS (sem alterar o original)."""
    return _strip(valor).upper()


def _converter_valor_br(raw: Any) -> Optional[float]:
    """
    Converte valores monetários brasileiros para float.

    Aceita:
        R$ 16.928,33  →  16928.33
        16.928,33     →  16928.33
        16928.33      →  16928.33
        245784        →  245784.0
        245784.0      →  245784.0

    Retorna None para:
        None, "", "X", "À DEFINIR", strings que não convertam
        Erros do Excel (#VALUE!, etc.)
    """
    if raw is None:
        return None

    # Valores numéricos diretos do Excel
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and (raw != raw):  # NaN
            return None
        return float(raw)

    # String
    s = str(raw).strip()
    if not s or s.upper() in {"X", "À DEFINIR", "A DEFINIR", "#VALUE!", "#REF!", "#N/A",
                               "#NAME?", "#DIV/0!", "#NULL!", "ERRO"}:
        return None

    # Remover símbolo de moeda e espaços iniciais/finais
    s = s.replace("R$", "").strip()

    # Extrair prefixo numérico de strings anotacionais como "2397,84 (?)"
    # Ex: "2397,84 (?)" → "2397,84"
    m_prefix = re.match(r"^([\d\.,]+)\s*[\(\[].+", s)
    if m_prefix:
        s = m_prefix.group(1).strip()

    # Remover espaços restantes (separadores de milhar com espaço)
    s = s.replace(" ", "")

    # Formato BR: ponto como separador de milhar, vírgula como decimal
    if re.match(r"^\d{1,3}(\.\d{3})*(,\d+)?$", s):
        s = s.replace(".", "").replace(",", ".")
    # Formato com vírgula mas sem ponto de milhar: 1234,56
    elif re.match(r"^\d+(,\d+)$", s):
        s = s.replace(",", ".")
    # Ponto como decimal (padrão): 16928.33
    elif re.match(r"^\d+(\.\d+)?$", s):
        pass
    else:
        return None

    try:
        return float(s)
    except ValueError:
        return None


def _localizar_colunas(ws, linha_header: int) -> Dict[str, Optional[int]]:
    """
    Lê a linha de cabeçalho e retorna um dict {campo: indice_coluna_0based}.
    
    Garante que TIPO_CONTRATO aponte para a PRIMEIRA ocorrência da coluna.
    Usa comparação case-insensitive + strip.
    """
    colunas: Dict[str, Optional[int]] = {
        "qtd":           None,
        "tipo_contrato": None,   # primeira ocorrência
        "valor_total":   None,
        "status":        None,
        "fonte_lead":    None,
        "vendedor":      None,
        "qtd_func":      None,   # QUANTIDADE DE FUNCIONARIOS
    }

    row = list(ws.iter_rows(min_row=linha_header, max_row=linha_header,
                             values_only=True))[0]

    for idx, val in enumerate(row):
        if val is None:
            continue
        nome = str(val).strip().upper()

        if nome == HEADER_QTD and colunas["qtd"] is None:
            colunas["qtd"] = idx

        elif nome == HEADER_TIPO_CONTRATO and colunas["tipo_contrato"] is None:
            colunas["tipo_contrato"] = idx

        elif nome == HEADER_VALOR_TOTAL and colunas["valor_total"] is None:
            colunas["valor_total"] = idx

        elif nome == HEADER_STATUS and colunas["status"] is None:
            colunas["status"] = idx

        # "FONTE DO LEAD" — planilha pode ter "FONTE DO LEAD" sem acento
        elif nome in {"FONTE DO LEAD", "FONTE DE LEAD"} and colunas["fonte_lead"] is None:
            colunas["fonte_lead"] = idx

        elif nome == HEADER_VENDEDOR and colunas["vendedor"] is None:
            colunas["vendedor"] = idx

        elif nome in {HEADER_QTD_FUNC, "QUANTIDADE DE FUNCIONÁRIOS"} and colunas["qtd_func"] is None:
            colunas["qtd_func"] = idx

    return colunas


# Textos que identificam linhas de sub-cabeçalho repetido no meio da planilha
_TEXTOS_SUBHEADER = {
    "EMPRESA", "CNPJ", "TIPO DE CONTRATO", "VALOR TOTAL",
    "STATUS", "VENDEDOR", "FONTE DO LEAD", "FONTE DO CONTATO",
    "SITUAÇAO DO CONTRATO", "OBSERVAÇÃO", "ENVIO", "FECHAMENTO",
}


def _linha_eh_subheader(row_vals: tuple, colunas: Dict[str, Optional[int]]) -> bool:
    """
    Detecta linhas que são sub-cabeçalhos repetidos no meio da planilha.
    Critério: coluna tipo_contrato ou status contém texto de cabeçalho conhecido,
    E a coluna QTD está vazia (sub-headers não têm número sequencial).
    """
    # QTD deve estar vazio em um sub-header
    idx_qtd = colunas.get("qtd")
    if idx_qtd is not None and idx_qtd < len(row_vals):
        qtd_val = row_vals[idx_qtd]
        if qtd_val is not None and str(qtd_val).strip():
            return False  # Tem QTD → não é sub-header

    # Verificar se campos de conteúdo contêm texto de cabeçalho
    for campo in ["tipo_contrato", "status", "fonte_lead", "vendedor"]:
        idx = colunas.get(campo)
        if idx is not None and idx < len(row_vals):
            val = row_vals[idx]
            if val is not None:
                txt = str(val).strip().upper()
                if txt in _TEXTOS_SUBHEADER:
                    return True
    return False


def _linha_tem_dado_comercial(row_vals: tuple, colunas: Dict[str, Optional[int]]) -> bool:
    """
    Verifica se a linha tem pelo menos um campo comercial não-vazio
    ALÉM do QTD. Evita importar linhas que só têm o número sequencial
    (ex: linha de rodapé/contador sem dados comerciais).
    """
    campos_comerciais = ["tipo_contrato", "status", "vendedor", "fonte_lead"]
    for campo in campos_comerciais:
        idx = colunas.get(campo)
        if idx is not None and idx < len(row_vals):
            val = row_vals[idx]
            if val is not None and str(val).strip():
                return True
    return False


# ---------------------------------------------------------------------------
# IMPORTADOR PRINCIPAL
# ---------------------------------------------------------------------------

def importar_excel(
    arquivo: Path = ARQUIVO_EXCEL,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Importa todas as abas do Excel e retorna:
        (registros, problemas_qualidade)

    registros: lista de dicts com o modelo normalizado
    problemas_qualidade: lista de dicts descrevendo anomalias encontradas
    """
    logger.info("Abrindo arquivo: %s", arquivo)
    wb = openpyxl.load_workbook(str(arquivo), data_only=True)

    registros: List[Dict[str, Any]] = []
    problemas: List[Dict[str, Any]] = []
    data_importacao = datetime.now().isoformat()
    contador_global = 0

    for aba_nome, (mes_num, mes_nome, ano) in ABA_PARA_MES.items():
        if aba_nome not in wb.sheetnames:
            logger.warning("Aba '%s' não encontrada no arquivo!", aba_nome)
            continue

        ws = wb[aba_nome]
        logger.info("Processando aba: %s", aba_nome)

        # Localizar colunas pelo texto do cabeçalho
        colunas = _localizar_colunas(ws, LINHA_HEADER)
        logger.info("  Mapeamento de colunas: %s", colunas)

        # Verificar se colunas obrigatórias foram encontradas
        obrigatorias = ["qtd", "tipo_contrato", "valor_total", "status"]
        for campo in obrigatorias:
            if colunas[campo] is None:
                logger.warning("  AVISO: coluna '%s' não encontrada em '%s'", campo, aba_nome)
                problemas.append({
                    "aba": aba_nome,
                    "linha": LINHA_HEADER,
                    "tipo": "COLUNA_AUSENTE",
                    "campo": campo,
                    "descricao": f"Coluna '{campo}' não localizada no cabeçalho da aba {aba_nome}",
                })

        contador_aba = 0

        for linha_num, row in enumerate(
            ws.iter_rows(min_row=LINHA_HEADER + 1, values_only=True),
            start=LINHA_HEADER + 1,
        ):
            # Pular linhas completamente vazias
            if all(v is None for v in row):
                continue

            # Filtrar sub-cabeçalhos repetidos no meio da planilha
            if _linha_eh_subheader(row, colunas):
                logger.debug("  Linha %d ignorada: sub-cabeçalho repetido", linha_num)
                problemas.append({
                    "aba":     aba_nome,
                    "linha":   linha_num,
                    "tipo":    "SUBHEADER_IGNORADO",
                    "campo":   "-",
                    "descricao": "Linha de sub-cabeçalho repetido ignorada",
                })
                continue

            # Verificar se é uma linha comercial válida
            if not _linha_tem_dado_comercial(row, colunas):
                continue

            # ---------------------------------------------------------------
            # Extrair campos brutos
            # ---------------------------------------------------------------
            def _get(campo: str) -> Any:
                idx = colunas.get(campo)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            qtd_raw        = _get("qtd")
            contrato_raw   = _get("tipo_contrato")
            valor_raw      = _get("valor_total")
            status_raw     = _get("status")
            fonte_raw      = _get("fonte_lead")
            vendedor_raw   = _get("vendedor")
            qtd_func_raw   = _get("qtd_func")

            # ---------------------------------------------------------------
            # Tratamentos
            # ---------------------------------------------------------------

            # QTD: preservar original
            qtd_original = qtd_raw

            # Tipo de Contrato: trim apenas
            tipo_contrato = _strip(contrato_raw) or None

            # Status: trim para armazenamento, normalizar apenas para comparação
            status_armazenado = _strip(status_raw) or None

            # Fonte do Lead: preservar texto original (apenas trim)
            fonte_lead = _strip(fonte_raw) or None

            # Vendedor: preservar texto original (apenas trim)
            vendedor = _strip(vendedor_raw) or None

            # Quantidade de funcionários: converter para int ou None
            qtd_func_convertida = converter_qtd_funcionarios(qtd_func_raw)

            # Curva ABC: classificar conforme regras de porte
            curva_abc = classificar_curva(qtd_func_convertida)

            # Valor total: converter formato BR
            valor_convertido = _converter_valor_br(valor_raw)
            flag_valor_invalido = False

            if valor_convertido is None and valor_raw is not None:
                s = str(valor_raw).strip()
                if s and s.upper() not in {"", "NONE"}:
                    flag_valor_invalido = True
                    problemas.append({
                        "aba":        aba_nome,
                        "linha":      linha_num,
                        "tipo":       "VALOR_INVALIDO",
                        "campo":      "valor_total",
                        "valor_raw":  str(valor_raw),
                        "descricao":  f"Valor '{valor_raw}' não pôde ser convertido para numérico",
                    })

            # ---------------------------------------------------------------
            # Montar registro normalizado (SEM dados pessoais)
            # ---------------------------------------------------------------
            contador_global += 1
            contador_aba    += 1

            registro: Dict[str, Any] = {
                "id_registro":        str(uuid.uuid4()),
                "mes_numero":         mes_num,
                "mes_nome":           mes_nome,
                "ano":                ano,
                "aba_origem":         aba_nome,
                "linha_origem":       linha_num,
                "qtd_original":       qtd_original,
                "tipo_contrato":      tipo_contrato,
                "valor_total":        valor_convertido,
                "status":             status_armazenado,
                "fonte_lead":         fonte_lead,
                "vendedor":           vendedor,
                "flag_valor_invalido":        flag_valor_invalido,
                # ── Campos Curva ABC (Etapa 2 — Qualidade de Vendas) ──
                "quantidade_funcionarios_original": qtd_func_raw,
                "quantidade_funcionarios":          qtd_func_convertida,
                "curva_abc_cliente":                curva_abc,
                "data_importacao":    data_importacao,
            }

            registros.append(registro)

        logger.info("  Registros importados: %d", contador_aba)

    wb.close()
    logger.info("Total importado: %d registros | %d problemas de qualidade",
                len(registros), len(problemas))
    return registros, problemas


# ---------------------------------------------------------------------------
# PERSISTÊNCIA
# ---------------------------------------------------------------------------

def salvar_base(registros: List[Dict[str, Any]], caminho: Path = SAIDA_JSON) -> None:
    """Persiste a base analítica em JSON."""
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(registros, f, ensure_ascii=False, indent=2, default=str)
    logger.info("Base analítica salva em: %s", caminho)


def salvar_qualidade(problemas: List[Dict[str, Any]], caminho: Path = LOG_QUALIDADE) -> None:
    """Persiste o relatório de qualidade em JSON."""
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(problemas, f, ensure_ascii=False, indent=2, default=str)
    logger.info("Relatório de qualidade salvo em: %s", caminho)


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    registros, problemas = importar_excel()
    salvar_base(registros)
    salvar_qualidade(problemas)
    print(f"\n✅ Importação concluída: {len(registros)} registros, {len(problemas)} problemas.")
